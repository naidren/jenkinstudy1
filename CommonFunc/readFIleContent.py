# -*- coding: utf-8 -*-
# ---
# @Software: PyCharm
# @File: readFIleContent.py
# @Author: chenzw
# @Institution: Guangzhou
# @E-mail: 2648738760@qq.com
# @Site: 读写文件，包括yaml，text,word,excel,pd,csv
# @Time: 1月 19, 2022
# ---
import openpyxl,yaml
from configparser import ConfigParser
import csv
class CsvHelp:
    """
    csv也可以通过字典格式写入或读取，有空再拓展。csv.DictWriter(),csv.DictReader(f)。
    写入一行：csv.writer(),
    """
    def get_line(self, csv_file):
        """
        读取csv文件的所有数据，返回嵌套列表，内层列表为每一行的数据。
        :param csv_file:csv文件的路径
        :return: type:list
        """
        with open(csv_file, mode='r', encoding='utf8') as csv_file:
            csv_data = csv.reader(csv_file)
            all_data = list()
            for ele in csv_data:
                col_data = list()
                for col in ele:
                    col_data.append(col)
                col_data = list()
                all_data.append(ele)
        return all_data

class yamlHelp(object):
    def yaml_read(self,file_path):
        """
        读取yaml文件的内容，读取出来的是dict类型.yaml.safe_load(f)中传的是数据流，不是文件路径。
        :param file_path: yaml文件路径
        :return: yaml文件内的数据，dict格式
        """
        with open(file=file_path,encoding='utf-8') as f:
            yaml_content = yaml.safe_load(f)
            # yaml.safe_load(f)
            return yaml_content

    def yaml_write(self,file_path,data):
        """
        写入单组数据到yaml文件。多组的时候，用dump_all().
        :param file_path: 写入文件的路径
        :param data: 写入yaml的文件。可以是dict,tuple,list,str,int等。
                    "a+"在尾部追加写入；”w+“清空原有内容重新写入。有个“+”是文件路径不存在时将自动创建。
        :return: None
        """
        with open(file_path, "w+", encoding='utf-8') as f:
            # allow_unicode是防止输入中文乱码
            yaml.dump(data, stream=f, allow_unicode=True)

class TextHelp(object):
    def read_one_line(self,path):
        """
        以行的形式读取，返回str
        """
        with open(file=path,mode='r',encoding='utf-8') as f:
            content = f.read()
            return content

    def read_multi_lines(self,path):
        """
        读取出所有行，返回list:[line1,line2]
        """
        with open(file=path,mode='r',encoding='utf-8') as f:
            content = f.readlines()
            return content

    def read_all_file(self,path):
        """
        读取text文本所有内容,返回str
        """
        with open(path, 'r', encoding="utf-8") as f:
            content = f.read()
            return content

class ExcelHelp:
    def __init__(self,excel_file):
        self.excel_file = excel_file
        self.wb = openpyxl.load_workbook(excel_file,read_only=False)

    def read_cell_value(self,sheetname,row:int,col):
        """
        读取一个单元格的数据
        :param sheetname:
        :param row:
        :param col:
        :return:
        """
        ws = self.wb[sheetname]
        if type(col) is int:
            if col>0 and row>0:
                cell_value = ws.cell(row=row,column=col).value
                return cell_value

        # 判断是col必须是个数为1且为字母的str.
        elif type(col) is str and len(col)==1 and col.isalpha():

            col_upper = col.upper()
            #a = ord("A") -->65,所以要减去64，A才能对应为1.
            col = ord(col_upper)-64
            cell_value = ws.cell(row=row, column=col).value
            return cell_value
        else:
            raise ValueError("row必须是大于0的数字或者一个字母！")


    def read_one_sheet_data(self,sheetname,ignore_first_row:bool=True):
        """
        读取一个sheet表的所有数据，并按照每行一个列表的格式输出。
        :param excel_file: excel文件的路径
        :param sheetname: sheet表名
        :param ignore_title: 是否保留标题，False为不保留，True为保留
        :return:返回两层嵌套列表，最外层为每行数据，里面一层为每个单元格的数据。
        """
        # wb = openpyxl.load_workbook(excel_file)
        sheet = self.wb[sheetname]
        sheet_data = list()
        row_data = list()
        for row in sheet:
            for cell in row:
                # 如果表格没值，则写入“”。
                if cell.value:
                    cell_value = cell.value
                else:
                    cell_value = ""
                row_data.append(cell_value)
            sheet_data.append(row_data)
            row_data = []
        if not ignore_first_row:
            return sheet_data
        else:
            try:
                sheet_data.remove(sheet_data[0])
                return sheet_data
            except:
                print("此表格：<{0}>没有数据".format(sheetname))
                return sheet_data

    def read_all_sheet_data(self,ignore_first_row:bool=False):
        """
        读取一个excel文件的所有sheet表的数据，并返回一个3层嵌套列表。
        :param ignore_title: 忽略读取标题
        :return:
        """
        sheets = self.wb.sheetnames
        all_sheet_datas = list()
        for son_sheet in sheets:
            all_sheet_datas.append(self.read_one_sheet_data(son_sheet,ignore_first_row))
        return all_sheet_datas
    def read_one_col_datas(self,sheetname,col:int,ignore_first_row:bool=True):
        """
        获取一个sheet表的col列数据，并以列表的形式返回。
        :param sheetname: 表名
        :param col: 列数，大于1
        :param ignore_first_row: 忽略首行（标题）
        :return: Type->list,某列的所有值。
        """
        ws = self.wb[sheetname]
        max_row = ws.max_row
        col_data = list()
        if ignore_first_row:
            start_row = 2
        else:
            start_row = 1
        if col>1:
            for i in range(start_row, max_row + 1):
                cell_data = ws.cell(row=i, column=col).value
                if cell_data is not None:
                    col_data.append(cell_data)
                else:
                    col_data.append('')
            return col_data
        else:
            raise IndexError("row必须是大于0的整数!")

    def write_data_in_cell(self,sheetname,row,col,content):
        ws = self.wb[sheetname]
        if type(col) is int:
            if col > 0 and row > 0:
                ws.cell(row=row, column=col).value=content

        # 判断是col必须是个数为1且为字母的str.
        elif type(col) is str and len(col) == 1 and col.isalpha():
            col_upper = col.upper()
            # a = ord("A") -->65,所以要减去64，A才能对应为1.
            col = ord(col_upper) - 64
            ws.cell(row=row, column=col).value = content
        else:
            raise ValueError("row必须是大于0的数字或者一个字母！")
        self.wb.save(self.excel_file)

class Wordhelp:
    pass

class PandasHelp:
    pass

class ConfHelp:

    def __init__(self,conf_file):
        # from configparser import ConfigParser
        self.conf_file = conf_file
        self.config = ConfigParser()
        self.config.read(filenames=self.conf_file,encoding="utf-8")

    def get_sections(self):
        # 获取配置文件中的所有section
        return self.config.sections()

    def get_items(self,section_name):
        # 获取指定的section中的所有配置选项名，即keys. 还有对应的value值
        return self.config.options(section_name)

    def get_one_item(self,section,item_key,value_type='str'):
        """
        获取一个section配置中的指定值。
        :param section: section名称
        :param item_key: item的key值
        :param value_type:返回值的类型，可以是str,int,bool,float
        :return:item_key对应的value值，类型是和value_type相同的类型。
        """
        if value_type == 'str':
            return self.config.get(section=section, option=item_key)
        elif value_type == 'int':
            return self.config.getint(section=section, option=item_key)
        elif value_type == 'float':
            return self.config.getfloat(section=section, option=item_key)
        elif value_type == 'bool':
            return self.config.getboolean(section=section, option=item_key)
        else:
            raise Exception('输入的value_type不正确：%s'%value_type)

    def add_in_conf(self,add_section,add_key,add_value,mode='w'):
        """
        新增一个section,并添加配置项和值.此section不能是已存在的section。
        :param add_section: 新增的section名称
        :param add_key: 选项名称
        :param add_value: 选项值
        :return: None
        """
        if not self.config.has_section(add_section):
            self.config.add_section(add_section)
            self.config.set(add_section,add_key,add_value)
            with open(self.conf_file, mode, encoding='utf-8') as f:
                self.config.write(f)
        else:
            raise Exception("错误!!!此section名已存在不能再新增：%s" % add_section)


    def edit_conf_by_key(self,section,key,edit_content):
        """
        新增或编辑配置项的值,传参都是str。值存在时会覆盖，即编辑。不存在则是新增。
        :param section:
        :param key:
        :param edit_content:
        :return:None
        """
        self.config[section][key] = edit_content
        with open(self.conf_file, 'w', encoding='utf-8') as f:
            self.config.write(f)

    def edit_conf_by_dict(self,section,**content):
        """
        编辑配置项的值,传参是字典。
        :param section:str,section名
        :param content:dict,{配置项名:配置项的值}
        :return: None
        """
        self.config[section] = content
        with open(self.conf_file, 'w', encoding='utf-8') as f:
            self.config.write(f)
    def delete_conf(self,section,key):
        self.config.remove_option(section=section,option=key)
        with open(self.conf_file, 'w', encoding='utf-8') as f:
            self.config.write(f)

    def delete_section(self,section):
        self.config.remove_section(section=section)
        with open(self.conf_file, 'w', encoding='utf-8') as f:
            self.config.write(f)

class ParseDataHelp:
    def parse_dict(self,data,key):
        pass
def count(d):
    # 获取嵌套字典的层数
    return max(count(v) if isinstance(v,dict) else 0 for v in d.values()) + 1

def dict_test(d):

    # for k,v in d.values():
    for k,v in d.items():
        # print(v)
        if isinstance(v,dict):
            # return max(count(v))+1
            print(dict_test(v))

            print(k,v)
        else:
            return 0

if __name__ == '__main__':
    data1 = {
        "writedata":{"line1":2,"line2":"牛牛"},
        "writedata2":{"line1":{"third":{"forth":"第四层数据"}},"line2":"牛牛"},
        "report":"file",
        "tuple":("3",5,1),
        "list":[2,4,1],
    }
    print(dict_test(data1))
    # print(data1.get("writedata2").get("line1"))
    # tuple1 = (8,8,8)
    # list1 = [9,9,9]
    # data2 = {"data":tuple1}
    # yaml_file=r'F:\projectcodes\PycharmProjects\static\test1.yaml'
    # yaml_file2=r'F:\projectcodes\PycharmProjects\static\testwrite.yaml'
    # conf_file = r"F:\projectcodes\PycharmProjects\static\setting.ini"
    # con = ConfHelp(conf_file=conf_file)
    # # con.delete_conf('lianxi','op3')
    # con.add_in_conf('lianxi3','add3','text3')
    # 获取line2的值
    # w_key = 'line2'
    # if w_key in data1.keys():
    #     print(data1.get(w_key))
    # else:
    #     for k,v in data1.items():
    #         # print(k,v)
    #         if type(v)==dict and w_key in v.keys():
    #             print(v.get(w_key))
    #         else:
    #             pass

    # txt_help = TextHelp()
    # txt_path = r"C:\\Users\\86166\\Desktop\\每日任务.txt"
    # content = txt_help.read_all_file(txt_path)
    # print(content)
